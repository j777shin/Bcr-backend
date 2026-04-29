"""
Load reference data from data/ref/*.xlsm into the database.

Populates:
  - countries: stub rows (country_code + country_name) for countries not yet seeded
  - country_metrics: ecosystem extent, loss rate, restorable area rows per country

Never overwrites rows that already exist — curated data (e.g. IDN) is preserved.
Usage: python load_ref_data.py
"""
import os

import openpyxl

from app import create_app, db
from app.models.global_layer import Country
from app.models.country_layer import CountryMetric

REF_DIR = os.path.join(os.path.dirname(__file__), 'data', 'ref')
XLSM = os.path.join(REF_DIR, 'Carbon-Cost Data Upload.xlsm')


def _fmt_ha(ha):
    if ha is None or ha == 0:
        return None
    if ha >= 1_000_000:
        return f"{ha / 1_000_000:.2f}M ha"
    if ha >= 1_000:
        return f"{ha / 1_000:.0f}K ha"
    return f"{ha:,.0f} ha"


def _fmt_rate(rate):
    if rate is None:
        return None
    return f"{rate * 100:.2f}%/yr"


def _sheet_rows(wb, sheet_name, min_row=2):
    ws = wb[sheet_name]
    headers = [c for c in next(ws.iter_rows(max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if any(v is not None for v in row):
            yield dict(zip(headers, row))


def load_country_stubs(wb):
    """Add country_code + country_name for every country not already in the DB."""
    added = 0
    for row in _sheet_rows(wb, 'Countries'):
        code = row.get('country_code')
        name = row.get('country')
        if not code or not name:
            continue
        if not Country.query.filter_by(country_code=code).first():
            db.session.add(Country(country_code=code, country_name=name))
            added += 1
    return added


def _add_metric(country_code, metric_name, metric_value, metric_subtext):
    """Insert metric only if (country_code, metric_name) does not already exist."""
    if not metric_value:
        return 0
    exists = CountryMetric.query.filter_by(
        country_code=country_code, metric_name=metric_name
    ).first()
    if not exists:
        db.session.add(CountryMetric(
            country_code=country_code,
            metric_name=metric_name,
            metric_value=metric_value,
            metric_subtext=metric_subtext,
        ))
        return 1
    return 0


def load_ecosystem_extent(wb):
    added = 0
    for row in _sheet_rows(wb, 'Ecosystem extent'):
        code = row.get('Country code')
        eco = row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source extent') or 'Global Mangrove Watch / WCMC'
        added += _add_metric(code, f"{eco} extent",
                             _fmt_ha(row.get('Extent')), source)
    return added


def load_ecosystem_loss(wb):
    added = 0
    for row in _sheet_rows(wb, 'Ecosystem loss'):
        code = row.get('Country code')
        eco = row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source') or 'Global Mangrove Watch'
        added += _add_metric(code, f"{eco} loss rate",
                             _fmt_rate(row.get('Loss rate')), source)
    return added


def load_restorable_land(wb):
    added = 0
    for row in _sheet_rows(wb, 'Restorable land'):
        code = row.get('Country code')
        eco = row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source') or 'Mapping Ocean Wealth'
        added += _add_metric(code, f"{eco} restorable area",
                             _fmt_ha(row.get('Restorable land')), source)
    return added


def load_all():
    app = create_app()
    with app.app_context():
        wb = openpyxl.load_workbook(XLSM, read_only=True, keep_vba=True)

        n_countries = load_country_stubs(wb)
        db.session.flush()  # ensure stubs exist before FK-referencing metrics

        n_extent = load_ecosystem_extent(wb)
        n_loss = load_ecosystem_loss(wb)
        n_restorable = load_restorable_land(wb)

        wb.close()
        db.session.commit()

        print(f'  ✓ country stubs added:         {n_countries}')
        print(f'  ✓ ecosystem extent metrics:    {n_extent}')
        print(f'  ✓ ecosystem loss rate metrics: {n_loss}')
        print(f'  ✓ restorable area metrics:     {n_restorable}')
        print(f'  ✓ ref data loaded: {n_countries + n_extent + n_loss + n_restorable} total rows')


if __name__ == '__main__':
    load_all()
