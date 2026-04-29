"""
Extract data from data/ref/*.xlsm into numbered CSVs in data/.

Outputs:
  14_countries_ref.csv       — country_code, country_name, map_cx, map_cy
  15_country_metrics_ref.csv — country_code, metric_name, metric_value, metric_subtext

Coordinates use an equirectangular projection calibrated so that
IDN (lat=-2, lng=118) maps to cx=966, cy=268 (matching the existing IDN record).

Run this script to regenerate the CSVs whenever the source xlsm files change.
Then run load_data.py (or seed.py) to push the CSVs into the database.
"""
import csv
import os

import openpyxl

REF_DIR = os.path.join(os.path.dirname(__file__), 'data', 'ref')
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
XLSM = os.path.join(REF_DIR, 'Carbon-Cost Data Upload.xlsm')
CENTROIDS_CSV = os.path.join(REF_DIR, 'country_centroids.csv')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _project(lat, lng):
    """Equirectangular → SVG cx/cy for viewBox 0 0 1200 560."""
    cx = round((lng + 180) / 360 * 1200 - 27.3, 1)
    cy = round((90 - lat) / 180 * 560 - 18.2, 1)
    return cx, cy


def _fmt_ha(ha):
    if ha is None or ha == 0:
        return None
    if ha >= 1_000_000:
        return f"{ha / 1_000_000:.2f}M ha"
    if ha >= 1_000:
        return f"{ha / 1_000:.0f}K ha"
    return f"{ha:,.0f} ha"


def _fmt_rate(rate):
    if rate is None:
        return None
    return f"{rate * 100:.2f}%/yr"


def _sheet_rows(wb, sheet_name):
    ws = wb[sheet_name]
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            yield dict(zip(headers, row))


def _write_csv(filename, fieldnames, rows):
    path = os.path.join(DATA_DIR, filename)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


# ── Extractors ────────────────────────────────────────────────────────────────

def extract_countries(wb):
    """14_countries_ref.csv — one row per unique country_code."""
    centroids = {}
    with open(CENTROIDS_CSV, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            cx, cy = _project(float(row['lat']), float(row['lng']))
            centroids[row['country_code']] = (cx, cy)

    seen = {}
    for row in _sheet_rows(wb, 'Countries'):
        code = row.get('country_code')
        name = row.get('country')
        if code and name and code not in seen:
            cx, cy = centroids.get(code, (None, None))
            seen[code] = {'country_code': code, 'country_name': name,
                          'map_cx': cx, 'map_cy': cy}

    rows = list(seen.values())
    return _write_csv('14_countries_ref.csv',
                      ['country_code', 'country_name', 'map_cx', 'map_cy'], rows)


def extract_country_metrics(wb):
    """15_country_metrics_ref.csv — ecosystem extent, loss rate, restorable area."""
    rows = []

    for row in _sheet_rows(wb, 'Ecosystem extent'):
        code, eco = row.get('Country code'), row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source extent') or 'Global Mangrove Watch / WCMC'
        val = _fmt_ha(row.get('Extent'))
        if val:
            rows.append({'country_code': code, 'metric_name': f"{eco} extent",
                         'metric_value': val, 'metric_subtext': source})

    for row in _sheet_rows(wb, 'Ecosystem loss'):
        code, eco = row.get('Country code'), row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source') or 'Global Mangrove Watch'
        val = _fmt_rate(row.get('Loss rate'))
        if val:
            rows.append({'country_code': code, 'metric_name': f"{eco} loss rate",
                         'metric_value': val, 'metric_subtext': source})

    for row in _sheet_rows(wb, 'Restorable land'):
        code, eco = row.get('Country code'), row.get('Ecosystem')
        if not code or not eco:
            continue
        source = row.get('Source') or 'Mapping Ocean Wealth'
        val = _fmt_ha(row.get('Restorable land'))
        if val:
            rows.append({'country_code': code, 'metric_name': f"{eco} restorable area",
                         'metric_value': val, 'metric_subtext': source})

    return _write_csv('15_country_metrics_ref.csv',
                      ['country_code', 'metric_name', 'metric_value', 'metric_subtext'], rows)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(XLSM, read_only=True, keep_vba=True)
    n_countries = extract_countries(wb)
    n_metrics = extract_country_metrics(wb)
    wb.close()
    print(f'  ✓ 14_countries_ref.csv:        {n_countries} rows')
    print(f'  ✓ 15_country_metrics_ref.csv:  {n_metrics} rows')


if __name__ == '__main__':
    main()
