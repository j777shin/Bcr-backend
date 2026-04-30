"""
Extract project-level CAPEX/OPEX cost data from the Carbon-Cost workbook
(Projects sheet) into data/16_project_costs_ref.csv.

Usage: python extract_project_costs.py
"""
import csv
import os

import openpyxl

ROOT = os.path.dirname(__file__)
SRC = os.path.join(ROOT, 'data', 'ref', 'Carbon-Cost Data Upload.xlsm')
OUT = os.path.join(ROOT, 'data', '16_project_costs_ref.csv')
SHEET = 'Projects'


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=False)
    ws = wb[SHEET]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in values):
            continue
        if values[0] is None:
            continue
        rows.append(values)

    with open(OUT, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)

    print(f'  ✓ Wrote {len(rows)} project-cost rows to {os.path.relpath(OUT, ROOT)}')


if __name__ == '__main__':
    main()
